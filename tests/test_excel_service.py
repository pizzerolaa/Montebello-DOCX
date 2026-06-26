from __future__ import annotations

from openpyxl import load_workbook

from app.config import Settings
from app.models.equipment import Equipment
from app.models.extraction_issue import ExtractionIssue
from app.models.lot import Lot
from app.models.project import Project
from app.models.signature import Signature
from app.models.source_document import SourceDocument
from app.models.work_item import EquipmentWorkItem
from app.services.excel_service import derived_work_done, export_project_excel, project_type


def test_project_type_and_derived_work_done() -> None:
    equipment = Equipment(equipment_type="PACKAGE", sequence=1)
    equipment.work_items.append(EquipmentWorkItem(sequence=1, title="BANDA", description="SUSTITUIDA."))
    equipment.work_items.append(EquipmentWorkItem(sequence=2, title=None, description="LIMPIEZA GENERAL."))
    project = Project(name="Prueba")
    project.equipment.append(Equipment(equipment_type="MINISPLIT", sequence=1))
    project.equipment.append(equipment)

    assert project_type(project) == "MIXED"
    assert derived_work_done(equipment) == "BANDA: SUSTITUIDA.\nLIMPIEZA GENERAL."


def test_export_project_excel_creates_required_sheets_and_legacy_columns(workspace_tmp_path) -> None:
    settings = Settings(storage_root=workspace_tmp_path / "storage")
    project = Project(
        id="project-id",
        name="Proyecto Excel",
        location="Tapachula",
        service_date_raw="10 DE JUNIO DE 2026",
        order_number="ABC-123",
        center_name="Hospital General",
        contract_date_raw="01 DE JUNIO DE 2026",
    )
    lot = Lot(id="lot-id", project_id=project.id, lot_number="01")
    project.lots.append(lot)
    equipment = Equipment(
        id="equipment-id",
        project_id=project.id,
        lot=lot,
        sequence=1,
        equipment_type="MINISPLIT",
        zone="CONSULTORIO",
        brand="YORK",
        capacity="1 TON",
        serial="00123",
    )
    equipment.work_items.append(
        EquipmentWorkItem(
            equipment_id=equipment.id,
            sequence=1,
            title="CAPACITOR",
            description="SUSTITUIDO.",
            source_text="CAPACITOR SUSTITUIDO.",
            is_custom=True,
        )
    )
    project.equipment.append(equipment)
    project.signatures.append(Signature(project_id=project.id, role="RECEIVER", name="Persona", position="Cargo"))
    project.source_documents.append(SourceDocument(id="source-id", project_id=project.id, original_filename="fuente.docx", safe_filename="fuente.docx"))
    project.extraction_issues.append(
        ExtractionIssue(
            project_id=project.id,
            source_document_id="source-id",
            severity="WARNING",
            entity_type="Equipment",
            field_name="serial",
            message="Aviso de prueba.",
        )
    )

    result = export_project_excel(project, settings)

    workbook = load_workbook(result.stored_path)
    assert workbook.sheetnames == [
        "General",
        "Lots",
        "Equipment",
        "Signatures",
        "WorkCatalog",
        "EquipmentWork",
        "Review",
        "Work",
        "Report",
    ]
    equipment_sheet = workbook["Equipment"]
    assert [cell.value for cell in equipment_sheet[1]] == [
        "equipment_id",
        "lot_id",
        "number",
        "equipment_type",
        "zone",
        "brand",
        "capacity",
        "serial",
        "work_done",
    ]
    assert equipment_sheet["H2"].value == "00123"
    assert equipment_sheet["H2"].number_format == "@"
    assert "CAPACITOR: SUSTITUIDO." in equipment_sheet["I2"].value
    assert workbook["Review"]["E2"].value == "Aviso de prueba."
