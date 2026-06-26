from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.config import Settings
from app.models.equipment import Equipment
from app.models.project import Project


MISSING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


@dataclass(frozen=True)
class ExcelExportResult:
    filename: str
    stored_path: Path


def export_project_excel(project: Project, settings: Settings) -> ExcelExportResult:
    """Export reviewed project data to a backward-compatible Excel workbook."""
    export_dir = (settings.storage_root / "exports" / project.id).resolve()
    storage_root = settings.storage_root.resolve()
    if storage_root not in export_dir.parents:
        raise ValueError("La ruta de exportacion no es valida.")
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    filename = f"{_safe_stem(project.name)}_{timestamp}.xlsx"
    path = export_dir / filename

    workbook = Workbook()
    general = workbook.active
    general.title = "General"
    _write_general(general, project)
    _write_lots(workbook.create_sheet("Lots"), project)
    _write_equipment(workbook.create_sheet("Equipment"), project)
    _write_signatures(workbook.create_sheet("Signatures"), project)
    _write_work_catalog(workbook.create_sheet("WorkCatalog"), project)
    _write_equipment_work(workbook.create_sheet("EquipmentWork"), project)
    _write_review(workbook.create_sheet("Review"), project)
    _write_legacy_work(workbook.create_sheet("Work"), project)
    _write_report(workbook.create_sheet("Report"))

    for sheet in workbook.worksheets:
        _format_sheet(sheet)

    workbook.save(path)
    return ExcelExportResult(filename=filename, stored_path=path)


def project_type(project: Project) -> str:
    types = {equipment.equipment_type for equipment in project.equipment if equipment.equipment_type != "UNKNOWN"}
    if len(types) == 1:
        return next(iter(types))
    if len(types) > 1:
        return "MIXED"
    return "UNKNOWN"


def derived_work_done(equipment: Equipment) -> str:
    parts = []
    for item in sorted(equipment.work_items, key=lambda work_item: work_item.sequence):
        if item.title:
            parts.append(f"{item.title}: {item.description}")
        else:
            parts.append(item.description)
    return "\n".join(parts)


def _write_general(sheet: Worksheet, project: Project) -> None:
    rows = [
        ("location", project.location),
        ("date", project.service_date_raw or project.service_date),
        ("order_number", project.order_number),
        ("center_name", project.center_name),
        ("equipment_count", len(project.equipment)),
        ("contract_date", project.contract_date_raw or project.contract_date),
        ("project_type", project_type(project)),
        ("client_name", project.client_name),
        ("status", project.status),
    ]
    sheet.append(["key", "value"])
    for key, value in rows:
        sheet.append([key, value if value is not None else ""])


def _write_lots(sheet: Worksheet, project: Project) -> None:
    sheet.append(["lot_id", "lot_number", "description"])
    for lot in project.lots:
        sheet.append([lot.id, lot.lot_number, lot.description or ""])


def _write_equipment(sheet: Worksheet, project: Project) -> None:
    sheet.append(["equipment_id", "lot_id", "number", "equipment_type", "zone", "brand", "capacity", "serial", "work_done"])
    for equipment in sorted(project.equipment, key=lambda item: item.sequence):
        sheet.append(
            [
                equipment.id,
                equipment.lot_id or "",
                equipment.sequence,
                equipment.equipment_type,
                equipment.zone or "",
                equipment.brand or "",
                equipment.capacity or "",
                equipment.serial or "",
                derived_work_done(equipment),
            ]
        )
        serial_cell = sheet.cell(row=sheet.max_row, column=8)
        serial_cell.number_format = "@"


def _write_signatures(sheet: Worksheet, project: Project) -> None:
    sheet.append(["role", "name", "position", "organization"])
    for signature in sorted(project.signatures, key=lambda item: item.sequence):
        sheet.append([signature.role, signature.name or "", signature.position or "", signature.organization or ""])


def _write_work_catalog(sheet: Worksheet, project: Project) -> None:
    sheet.append(["work_id", "equipment_type", "component", "default_description"])
    seen = set()
    for equipment in project.equipment:
        for work_item in equipment.work_items:
            if work_item.catalog_item and work_item.catalog_item.id not in seen:
                seen.add(work_item.catalog_item.id)
                sheet.append(
                    [
                        work_item.catalog_item.id,
                        work_item.catalog_item.equipment_type,
                        work_item.catalog_item.component,
                        work_item.catalog_item.default_description,
                    ]
                )


def _write_equipment_work(sheet: Worksheet, project: Project) -> None:
    sheet.append(["equipment_id", "sequence", "work_id", "title", "description", "is_custom"])
    for equipment in sorted(project.equipment, key=lambda item: item.sequence):
        for work_item in sorted(equipment.work_items, key=lambda item: item.sequence):
            sheet.append(
                [
                    equipment.id,
                    work_item.sequence,
                    work_item.catalog_item_id or "",
                    work_item.title or "",
                    work_item.description,
                    work_item.is_custom,
                ]
            )


def _write_review(sheet: Worksheet, project: Project) -> None:
    sheet.append(["severity", "source_document", "entity", "field", "message", "resolved"])
    documents = {document.id: document.original_filename for document in project.source_documents}
    for issue in project.extraction_issues:
        sheet.append(
            [
                issue.severity,
                documents.get(issue.source_document_id or "", ""),
                issue.entity_type or "",
                issue.field_name or "",
                issue.message,
                issue.resolved,
            ]
        )


def _write_legacy_work(sheet: Worksheet, project: Project) -> None:
    sheet.append(["work"])
    seen = set()
    for equipment in project.equipment:
        for work_item in sorted(equipment.work_items, key=lambda item: item.sequence):
            value = f"{work_item.title}: {work_item.description}" if work_item.title else work_item.description
            if value not in seen:
                seen.add(value)
                sheet.append([value])


def _write_report(sheet: Worksheet) -> None:
    sheet.append(["Report"])
    sheet.append(["Reporte fotografico fuera de alcance del MVP."])


def _format_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    max_column = sheet.max_column
    max_row = sheet.max_row
    if max_column:
        sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value == "":
                cell.fill = MISSING_FILL
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 45)


def _safe_stem(value: str) -> str:
    cleaned = "".join(char if char.isalnum() or char in {"-", "_"} else "_" for char in value.strip())
    return cleaned.strip("_") or "project"
